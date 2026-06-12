"""
Populate RELISH_Entity_Import_Template.xlsx with Payee data from Relish Approvals Supabase DB.
Reads ONLY from Supabase. Writes ONLY to the Excel file.
Appends real payee rows after the existing sample rows, preserving headers + samples.

Usage:
  python populate_entities_from_approvals.py
  python populate_entities_from_approvals.py <supabase_url> <service_key>
"""

import json
import sys
import os
import urllib.request
import urllib.error
import urllib.parse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from copy import copy

# ── Supabase credentials ───────────────────────────────────────────────────────
# Priority: CLI args > environment variables > hardcoded fallback
if len(sys.argv) == 3:
    SUPABASE_URL = sys.argv[1]
    SUPABASE_SERVICE_KEY = sys.argv[2]
else:
    SUPABASE_URL = os.environ.get("SUPABASE_URL", "https://ewbguvwrejdvlhzcqlbp.supabase.co")
    SUPABASE_SERVICE_KEY = os.environ.get("SUPABASE_SERVICE_KEY", "PASTE_SERVICE_KEY_HERE")

EXCEL_PATH = r"c:\Users\user\Desktop\SOFTWARE DEV\APP Folders\relish-business-suite\tools\RELISH_Entity_Import_Template.xlsx"


def supabase_get(table, params=None):
    """Fetch rows from a Supabase table via REST API (read-only GET)."""
    url = f"{SUPABASE_URL}/rest/v1/{table}"
    if params:
        url += "?" + urllib.parse.urlencode(params)
    req = urllib.request.Request(url)
    req.add_header("apikey", SUPABASE_SERVICE_KEY)
    req.add_header("Authorization", f"Bearer {SUPABASE_SERVICE_KEY}")
    req.add_header("Accept", "application/json")
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            return json.loads(resp.read().decode())
    except urllib.error.HTTPError as e:
        body = e.read().decode()
        raise RuntimeError(f"HTTP {e.code} fetching {table}: {body}")


def derive_role(payee):
    """
    Best-effort role derivation from available payee fields.
      is_staff=True  → Staff
      Otherwise      → Vendor  (default; reviewer can adjust)
    """
    if payee.get("is_staff"):
        return "Staff"
    return "Vendor"


def map_payee_to_row(payee, company_map):
    """Map a Supabase payee record to the 28-column Entities sheet row."""
    company_id = payee.get("company_id", "")
    # Use the short company code if a mapping exists, else use the raw id
    company_code = company_map.get(company_id, company_id)

    role = derive_role(payee)
    name = payee.get("name") or ""
    alias = payee.get("alias") or None
    mobile = payee.get("mobile") or None
    bank_account = payee.get("bank_account") or None
    ifsc = payee.get("ifsc") or None
    upi_id = payee.get("upi_id") or None

    # Column order matches the Entities sheet header (28 columns):
    # role*, company_code*, display_name*, alias, mobile, mobile_alt, email,
    # gstin, pan, local_reg_number, local_tax_number,
    # address_line1, address_line2, city, state, pincode, country,
    # bank_name, bank_account_holder, bank_account_number, bank_ifsc, bank_swift,
    # upi_id, pramaana_ledger_name, designation, department, source, notes
    return [
        role,           # role *
        company_code,   # company_code *
        name,           # display_name *
        alias,          # alias
        mobile,         # mobile
        None,           # mobile_alt
        None,           # email
        None,           # gstin
        None,           # pan
        None,           # local_reg_number
        None,           # local_tax_number
        None,           # address_line1
        None,           # address_line2
        None,           # city
        None,           # state
        None,           # pincode
        None,           # country
        None,           # bank_name
        name if bank_account else None,  # bank_account_holder (use name if bank exists)
        bank_account,   # bank_account_number
        ifsc,           # bank_ifsc
        None,           # bank_swift
        upi_id,         # upi_id
        name,           # pramaana_ledger_name (default to name)
        None,           # designation
        None,           # department
        "Approvals",    # source
        None,           # notes
    ]


def main():
    print("── Relish Approvals → Entity Import Populator ──")

    # 1. Fetch companies (to resolve company_id → code)
    print("Fetching companies…")
    companies = supabase_get("companies", {"select": "id,name"})
    # company_id in payees is the TEXT primary key (often already the short code)
    company_map = {c["id"]: c["id"] for c in companies}  # id IS the code
    print(f"  Found {len(companies)} companies: {[c['id'] for c in companies]}")

    # 2. Fetch all payees
    print("Fetching payees…")
    payees = supabase_get(
        "payees",
        {"select": "id,company_id,name,alias,mobile,bank_account,ifsc,upi_id,is_staff",
         "order": "company_id.asc,name.asc"}
    )
    print(f"  Found {len(payees)} payees")

    if not payees:
        print("No payee records found — nothing to write.")
        return

    # 3. Open the Excel workbook
    print(f"Opening workbook: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb["Entities"]

    # Determine where existing data ends (find last non-empty row)
    last_row = ws.max_row
    # Walk backward to skip truly blank trailing rows
    while last_row > 1 and all(ws.cell(last_row, c).value is None for c in range(1, 29)):
        last_row -= 1

    start_row = last_row + 1
    print(f"  Existing data ends at row {last_row}; new rows start at {start_row}")

    # 4. Copy formatting from the last sample row (row 5 if it exists, else row 2)
    template_row_idx = min(5, last_row) if last_row >= 2 else 2

    # 5. Write payee rows
    written = 0
    for payee in payees:
        row_data = map_payee_to_row(payee, company_map)
        target_row = start_row + written

        for col_idx, value in enumerate(row_data, start=1):
            src_cell = ws.cell(template_row_idx, col_idx)
            tgt_cell = ws.cell(target_row, col_idx)
            tgt_cell.value = value
            # Copy basic font/fill/alignment from template row
            if src_cell.font:
                tgt_cell.font = copy(src_cell.font)
            if src_cell.fill and src_cell.fill.fill_type != "none":
                tgt_cell.fill = copy(src_cell.fill)
            if src_cell.alignment:
                tgt_cell.alignment = copy(src_cell.alignment)

        written += 1

    print(f"  Wrote {written} payee rows (rows {start_row}–{start_row + written - 1})")

    # 6. Save
    wb.save(EXCEL_PATH)
    print(f"Saved → {EXCEL_PATH}")
    print("Done.")


if __name__ == "__main__":
    main()
